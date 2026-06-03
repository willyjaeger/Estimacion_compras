from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import pandas as pd
import pdfplumber
import re
import io
import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import CellIsRule
import base64
import anthropic

app = Flask(__name__, static_folder="static", static_url_path="")
CORS(app)

ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")

EXCLUIR_PALABRAS = ["ETIQUETA", "BOLSA", "CAJA", "CINTA", "CARTON", "FRASCO", "BOBINA"]


def excel_serial_to_date(serial):
    """Convierte número serial de Excel a datetime."""
    try:
        serial = float(serial)
        return datetime(1899, 12, 30) + timedelta(days=serial)
    except Exception:
        return None


def procesar_fotos(imagenes_b64, tipos):
    """Usa Claude Vision para extraer stock físico de las fotos."""
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

    content = []
    for b64, tipo in zip(imagenes_b64, tipos):
        content.append({
            "type": "image",
            "source": {"type": "base64", "media_type": tipo, "data": b64}
        })

    content.append({
        "type": "text",
        "text": (
            "Estas imágenes son notas manuscritas con inventario de stock físico de materias primas. "
            "Extraé TODOS los productos y sus cantidades en kilogramos. "
            "Si hay sumas anotadas (ej: '84k + 15k'), calculá el total. "
            "Respondé ÚNICAMENTE con un JSON válido, sin texto adicional, sin markdown, sin backticks. "
            "Formato exacto: [{\"producto\": \"NOMBRE EN MAYUSCULAS\", \"kg\": 123.5}, ...]"
        )
    })

    resp = client.messages.create(
        model="claude-opus-4-5",
        max_tokens=2000,
        messages=[{"role": "user", "content": content}]
    )

    texto = resp.content[0].text.strip()
    texto = re.sub(r"```json|```", "", texto).strip()
    datos = __import__("json").loads(texto)
    df = pd.DataFrame(datos)
    df.columns = ["PRODUCTO", "STOCK FÍSICO (FOTOS)"]
    df["PRODUCTO"] = df["PRODUCTO"].str.upper().str.strip()
    return df


def procesar_pdf(pdf_bytes):
    """Extrae SKUs y cantidades del PDF con pdfplumber."""
    skus = {}
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            # Busca patrones: código numérico seguido de cantidad
            # Patrones comunes en documentos inbound: "04041  120", "SKU: 04041 Cant: 120"
            matches = re.findall(r"\b(\d{4,6})\b.*?(\d+(?:\.\d+)?)\s*(?:u|un|unid|kg|k)?\b", text, re.IGNORECASE)
            for sku, cant in matches:
                try:
                    skus[sku] = skus.get(sku, 0) + float(cant)
                except Exception:
                    pass
    return skus


def procesar_excel(xlsm_bytes):
    """Lee la hoja Combinada y calcula todos los consumos."""
    df = pd.read_excel(io.BytesIO(xlsm_bytes), sheet_name="Combinada", engine="openpyxl")

    # Convertir Fechaconv
    def parse_fecha(v):
        try:
            f = float(v)
            return excel_serial_to_date(f)
        except Exception:
            try:
                return pd.to_datetime(v, errors="coerce")
            except Exception:
                return None

    df["Fechaconv"] = df["Fechaconv"].apply(parse_fecha)

    # Filtrar insumos no alimenticios
    mask_excluir = df["Producto Componente"].astype(str).str.upper().apply(
        lambda x: any(p in x for p in EXCLUIR_PALABRAS)
    )
    df = df[~mask_excluir].copy()

    fecha_hoy = df["Fechaconv"].dropna().max()
    if pd.isna(fecha_hoy):
        fecha_hoy = datetime.today()
    fecha_30 = fecha_hoy - timedelta(days=30)

    # RAPPI FUTURO
    rappi = (
        df[(df["Base"] == "Rappi") & (df["Fechaconv"] >= fecha_hoy)]
        .groupby("Producto Componente")["Cantidad Consumida x KG"]
        .sum()
        .reset_index()
        .rename(columns={"Producto Componente": "PRODUCTO", "Cantidad Consumida x KG": "RAPPI (FUTURO)"})
    )

    # MELI 30D
    meli = (
        df[(df["Base"] == "MELI") & (df["Fechaconv"] >= fecha_30) & (df["Fechaconv"] <= fecha_hoy)]
        .groupby("Producto Componente")["Cantidad Consumida x KG"]
        .sum()
        .reset_index()
        .rename(columns={"Producto Componente": "PRODUCTO", "Cantidad Consumida x KG": "MERCADO LIBRE (30D)"})
    )

    # MISTER NUT 30D
    mrnut = (
        df[(df["Base"] == "Mister Nut") & (df["Fechaconv"] >= fecha_30) & (df["Fechaconv"] <= fecha_hoy)]
        .groupby("Producto Componente")["Cantidad Consumida x KG"]
        .sum()
        .reset_index()
        .rename(columns={"Producto Componente": "PRODUCTO", "Cantidad Consumida x KG": "MISTER NUT (30D)"})
    )

    # ÚLTIMO PROVEEDOR
    df_prov = df.dropna(subset=["Proveedor Componente"]).sort_values("Fechaconv")
    ultimo_prov = (
        df_prov.groupby("Producto Componente")["Proveedor Componente"]
        .last()
        .reset_index()
        .rename(columns={"Producto Componente": "PRODUCTO", "Proveedor Componente": "ÚLTIMO PROVEEDOR"})
    )

    return rappi, meli, mrnut, ultimo_prov, df


def planificado_full(skus_pdf, xlsm_bytes):
    """Cruza SKUs del PDF con recetas del Excel."""
    df = pd.read_excel(io.BytesIO(xlsm_bytes), sheet_name="Combinada", engine="openpyxl")
    mask_excluir = df["Producto Componente"].astype(str).str.upper().apply(
        lambda x: any(p in x for p in EXCLUIR_PALABRAS)
    )
    df = df[~mask_excluir].copy()
    df["Articulo"] = df["Articulo"].astype(str).str.strip()

    resultados = {}
    for sku, cant_unidades in skus_pdf.items():
        sub = df[df["Articulo"] == str(sku)].drop_duplicates(
            subset=["Producto Componente"]
        )
        for _, row in sub.iterrows():
            comp = str(row["Producto Componente"]).upper().strip()
            kg_por_unidad = float(row["Cantidad x Compuesto"]) if pd.notna(row["Cantidad x Compuesto"]) else 0
            resultados[comp] = resultados.get(comp, 0) + cant_unidades * kg_por_unidad

    df_plan = pd.DataFrame(list(resultados.items()), columns=["PRODUCTO", "PLANIFICADO FULL"])
    return df_plan


def generar_excel(df_final):
    """Genera el archivo Excel con estilos."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Control de Stock"

    columnas = [
        "PRODUCTO", "ÚLTIMO PROVEEDOR", "STOCK FÍSICO (FOTOS)",
        "RAPPI (FUTURO)", "STOCK REAL", "MERCADO LIBRE (30D)",
        "MISTER NUT (30D)", "PLANIFICADO FULL"
    ]

    # Cabecera
    header_fill = PatternFill("solid", fgColor="002060")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, col_name in enumerate(columnas, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # Datos
    for row_idx, row in df_final.iterrows():
        for col_idx, col_name in enumerate(columnas, 1):
            val = row.get(col_name, 0)
            ws.cell(row=row_idx + 2, column=col_idx, value=val)

    # Formato condicional STOCK REAL (columna E)
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    red_font = Font(color="9C0006", bold=True)
    stock_real_col = "E"
    last_row = len(df_final) + 1
    ws.conditional_formatting.add(
        f"{stock_real_col}2:{stock_real_col}{last_row}",
        CellIsRule(operator="lessThan", formula=["0"], fill=red_fill, font=red_font)
    )

    # Ancho de columnas automático
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@app.route("/")
def index():
    return app.send_static_file("index.html")


@app.route("/generar", methods=["POST"])
def generar():
    try:
        # Recibir archivos
        fotos = request.files.getlist("fotos")
        pdf_file = request.files.get("pdf")
        excel_file = request.files.get("excel")

        if not fotos or not pdf_file or not excel_file:
            return jsonify({"error": "Faltan archivos. Se necesitan fotos, PDF y Excel."}), 400

        # Procesar fotos con Claude Vision
        imagenes_b64 = []
        tipos = []
        for f in fotos:
            contenido = f.read()
            imagenes_b64.append(base64.b64encode(contenido).decode())
            tipos.append(f.content_type or "image/jpeg")

        df_stock = procesar_fotos(imagenes_b64, tipos)

        # Leer PDF y Excel
        pdf_bytes = pdf_file.read()
        xlsm_bytes = excel_file.read()

        skus_pdf = procesar_pdf(pdf_bytes)
        rappi, meli, mrnut, ultimo_prov, _ = procesar_excel(xlsm_bytes)
        df_plan = planificado_full(skus_pdf, xlsm_bytes)

        # Merge
        df = df_stock.copy()
        for df_merge in [rappi, meli, mrnut, df_plan, ultimo_prov]:
            df = df.merge(df_merge, on="PRODUCTO", how="outer")

        # Rellenar nulos numéricos
        num_cols = ["STOCK FÍSICO (FOTOS)", "RAPPI (FUTURO)", "MERCADO LIBRE (30D)", "MISTER NUT (30D)", "PLANIFICADO FULL"]
        for col in num_cols:
            if col not in df.columns:
                df[col] = 0
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

        df["ÚLTIMO PROVEEDOR"] = df["ÚLTIMO PROVEEDOR"].fillna("")
        df["STOCK REAL"] = df["STOCK FÍSICO (FOTOS)"] - df["RAPPI (FUTURO)"]

        # Redondear
        for col in num_cols + ["STOCK REAL"]:
            df[col] = df[col].round(2)

        # Filtrar filas con todo en cero
        df = df[
            (df["STOCK FÍSICO (FOTOS)"] != 0) |
            (df["RAPPI (FUTURO)"] != 0) |
            (df["MERCADO LIBRE (30D)"] != 0) |
            (df["MISTER NUT (30D)"] != 0) |
            (df["PLANIFICADO FULL"] != 0)
        ].reset_index(drop=True)

        # Generar Excel
        excel_output = generar_excel(df)

        return send_file(
            excel_output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="Reporte_Control_Stock_Actualizado.xlsx"
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
